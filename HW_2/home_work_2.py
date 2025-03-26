import random
import string
import csv
import pandas as pd
from openpyxl import load_workbook, Workbook
from flask import Flask

app = Flask(__name__)

@app.route("/")
def hello_world():
    return "<p>Hello, World!!</p>"


@app.route("/password")
def generate_password():
    password_lenght = random.randint(10,20)
    pass_symbols = string.digits + string.ascii_lowercase + string.ascii_lowercase + "!#$%&/()>=?_+-"
    password = ''.join(random.choice(pass_symbols) for _ in range(password_lenght))
    return password


@app.route("/calculate_average")
def calculate_average():
    hw_excel = Workbook()
    activate_excel_file = hw_excel.active
    with open("hw.csv", "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            activate_excel_file.append(row)
    hw_excel.save("hw.xlsx")
    #put_file = load_workbook("hw.csv")
    #activate_file = put_file.active
    values_height = [cell.value for cell in activate_excel_file["B"] if isinstance(cell.value, (int, float, complex))]
    string_values_height = str(values_height)
    #sum_values_height = sum(values_height)
    #print(values_height)

    return string_values_height


@app.route("/calculate_average_pandas")
def calculate_average_pandas():
    read_this_csv = pd.read_csv("hw.csv")
    read_this_csv.to_excel("hw1.xlsx", index=False, engine="openpyxl")
    read_this_excel = pd.read_excel("hw1.xlsx", engine="openpyxl")
    sum_val_height = read_this_excel[" Height(Inches)"].sum()
    string_sum_val_height = str(sum_val_height)
    sum_val_weight = read_this_excel[" Weight(Pounds)"].sum()
    string_sum_val_weight = str(sum_val_weight)
    return (f' The total height: {string_sum_val_height}\n'
            f' The total weight: {string_sum_val_weight}')



if __name__ =='__main__':
    app.run(
        'localhost', debug = True
    )